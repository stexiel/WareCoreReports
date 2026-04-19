import pandas as pd
import sys
from openpyxl import load_workbook

def view_excel(filepath):
    """Просмотр Excel файла в консоли"""
    try:
        # Читаем через pandas (показывает данные)
        df = pd.read_excel(filepath, header=None)
        print("\n" + "="*80)
        print(f"ФАЙЛ: {filepath}")
        print("="*80)
        print("\n--- ДАННЫЕ ---")
        print(df.to_string(index=False, header=False))
        
        # Читаем через openpyxl (показывает стили)
        wb = load_workbook(filepath)
        ws = wb.active
        print(f"\n--- ИНФО ---")
        print(f"Лист: {ws.title}")
        print(f"Строк: {ws.max_row}, Столбцов: {ws.max_column}")
        print(f"Диапазон: {ws.dimensions}")
        
    except Exception as e:
        print(f"Ошибка: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        view_excel(sys.argv[1])
    else:
        # По умолчанию открываем файл из папки
        view_excel(r"c:\Users\Administrator\Desktop\WareCore Reports\Лист Microsoft Excel.xlsx")
